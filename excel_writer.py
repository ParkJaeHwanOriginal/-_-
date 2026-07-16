import os
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
import config

def save_to_excel(results):
    """
    results: { keyword: [ { 'date': 'YYYY/MM/DD', 'provider': '언론사명', 'title': '제목', 'url': '링크' }, ... ] }
    형태의 5개 매체(무등 제외 4개) 수집 결과를 누적하여 엑셀에 저장합니다.
    """
    file_path = config.OUTPUT_FILE_PATH
    
    # 엑셀 로드 또는 생성
    if os.path.exists(file_path):
        wb = openpyxl.load_workbook(file_path)
    else:
        wb = openpyxl.Workbook()
        # 기본 시트 제거
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])
            
    # 시트 준비 및 기본 헤더 삽입
    if "일일요약" not in wb.sheetnames:
        ws_summary = wb.create_sheet("일일요약")
        ws_summary.append(["날짜", "키워드", "빅카인즈", "동아일보", "광주매일", "남도일보", "총합계"])
    else:
        ws_summary = wb["일일요약"]
        
    if "상세뉴스목록" not in wb.sheetnames:
        ws_detail = wb.create_sheet("상세뉴스목록")
        ws_detail.append(["수집일자", "기사일자", "언론사명", "검색어(키워드)", "기사제목", "링크 URL"])
    else:
        ws_detail = wb["상세뉴스목록"]
        
    today_str = datetime.now().strftime("%Y/%m/%d")
    
    # 1. 상세뉴스목록 데이터 적재 (중복 방지)
    # 기존 상세 목록의 (기사일자, 언론사명, 기사제목) 튜플 세트를 만들어 중복을 필터링합니다.
    existing_details = set()
    for row in list(ws_detail.iter_rows(values_only=True))[1:]:
        if len(row) >= 5 and row[1] and row[2] and row[4]:
            existing_details.add((str(row[1]).strip(), str(row[2]).strip(), str(row[4]).strip()))
            
    # 매체 이름 매핑용 헬퍼
    provider_map = {
        "bigkinds": config.MEDIA_SITES["bigkinds"]["name"],
        "donga": config.MEDIA_SITES["donga"]["name"],
        "kjdaily": config.MEDIA_SITES["kjdaily"]["name"],
        "namdonews": config.MEDIA_SITES["namdonews"]["name"]
    }
    
    # 키워드별/매체별 기사 건수 카운터 준비
    summary_counts = {kw: {p_name: 0 for p_name in provider_map.values()} for kw in config.SEARCH_KEYWORDS}
    
    for kw, articles in results.items():
        if kw not in summary_counts:
            summary_counts[kw] = {p_name: 0 for p_name in provider_map.values()}
            
        for art in articles:
            # origin_site를 사용하여 매핑 이름을 구하고, 없으면 provider 기본값 사용
            origin = art.get("origin_site")
            p_name = provider_map.get(origin, art["provider"])
            
            if p_name in summary_counts[kw]:
                summary_counts[kw][p_name] += 1
                
            # 상세뉴스목록 중복 체크 및 추가
            detail_key = (art["date"].strip(), p_name.strip(), art["title"].strip())
            if detail_key not in existing_details:
                ws_detail.append([
                    today_str,
                    art["date"],
                    p_name,
                    kw,
                    art["title"],
                    art["url"]
                ])
                existing_details.add(detail_key)
                
    # 2. 일일요약 데이터 적재 (중복 업데이트 방지)
    # 동일 날짜의 동일 키워드가 존재하면 값을 덮어쓰고, 없으면 새로 추가합니다.
    for kw in config.SEARCH_KEYWORDS:
        counts = summary_counts[kw]
        bigkinds_cnt = counts[provider_map["bigkinds"]]
        donga_cnt = counts[provider_map["donga"]]
        kjdaily_cnt = counts[provider_map["kjdaily"]]
        namdo_cnt = counts[provider_map["namdonews"]]
        total_cnt = bigkinds_cnt + donga_cnt + kjdaily_cnt + namdo_cnt
        
        row_updated = False
        # 1행은 헤더이므로 2행부터 순회합니다.
        for row_idx in range(2, ws_summary.max_row + 1):
            cell_date = ws_summary.cell(row=row_idx, column=1).value
            cell_kw = ws_summary.cell(row=row_idx, column=2).value
            if str(cell_date) == today_str and str(cell_kw) == kw:
                # 기존 행 데이터 갱신
                ws_summary.cell(row=row_idx, column=3, value=bigkinds_cnt)
                ws_summary.cell(row=row_idx, column=4, value=donga_cnt)
                ws_summary.cell(row=row_idx, column=5, value=kjdaily_cnt)
                ws_summary.cell(row=row_idx, column=6, value=namdo_cnt)
                ws_summary.cell(row=row_idx, column=7, value=total_cnt)
                row_updated = True
                break
                
        if not row_updated:
            ws_summary.append([
                today_str,
                kw,
                bigkinds_cnt,
                donga_cnt,
                kjdaily_cnt,
                namdo_cnt,
                total_cnt
            ])
            
    # 3. 디자인 스타일 서식 적용 (비즈니스 프리미엄 톤)
    navy_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
    white_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="맑은 고딕", size=10)
    
    thin_border = Border(
        left=Side(style='thin', color='D3D3D3'),
        right=Side(style='thin', color='D3D3D3'),
        top=Side(style='thin', color='D3D3D3'),
        bottom=Side(style='thin', color='D3D3D3')
    )
    
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    for ws in [ws_summary, ws_detail]:
        # 헤더 스타일
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = navy_fill
            cell.font = white_font
            cell.alignment = center_align
            cell.border = thin_border
            
        # 데이터 행 스타일
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = data_font
                cell.border = thin_border
                
                # 데이터 값에 따른 정렬 규칙 지정
                if ws == ws_summary:
                    if col_idx in [1, 2]:
                        cell.alignment = center_align
                    else:
                        cell.alignment = right_align
                        cell.number_format = '#,##0'  # 천단위 컴마
                else:
                    if col_idx in [1, 2, 3, 4]:
                        cell.alignment = center_align
                    else:
                        cell.alignment = left_align
                        
        # 열 너비 비례 자동 조절 (Auto-fit Columns)
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                # 한글 글자 바이트 가중치를 두어 너비 계산
                cell_len = len(val_str.encode('utf-8', errors='replace'))
                adjusted_len = cell_len * 0.45 + 3
                if adjusted_len > max_len:
                    max_len = adjusted_len
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            # URL 등이 과도하게 길어져 엑셀 창이 벌어지는 것을 제한 (최대 55너비)
            ws.column_dimensions[col_letter].width = min(max_len, 55)
            
    # 엑셀 파일 저장
    wb.save(file_path)
    print(f"-> 엑셀 결과 보고서 파일이 성공적으로 갱신되었습니다: {file_path}")
